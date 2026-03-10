"""
Requirements Excel Generator
Creates simple and advanced requirements tracking templates with UUIDs

Usage:
    python requirements-excel-generator.py [output_directory]
    
Generates:
    - REQUIREMENTS-SIMPLE.xlsx: 4 columns (Feature, Approved, Notes, UUID)
    - REQUIREMENTS-ADVANCED.xlsx: 13 columns with feature groups, tags, versions, decisions
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import uuid
import sys
import os

def gen_uuid():
    """Generate 6-character hex UUID"""
    return uuid.uuid4().hex[:6]

def create_simple_template(output_dir):
    """Create simple requirements template optimized for keyboard navigation"""
    wb = Workbook()
    wb.remove(wb.active)
    
    header_fill = PatternFill(start_color='2F5496', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    ws = wb.create_sheet('Requirements')
    
    # Column order optimized for Tab navigation: Feature, Approved, Notes, UUID
    headers = ['Feature', 'Approved', 'Notes', 'UUID']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sample data
    sample_data = [
        ['Sample feature 1', 'Y', 'Approved as-is', gen_uuid()],
        ['Sample feature 2', 'With Changes', 'Needs minor tweaks', gen_uuid()],
        ['Sample feature 3', 'N', 'Out of scope', gen_uuid()],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Color code Approved column
            if col_idx == 2:
                if value == 'Y':
                    cell.fill = PatternFill(start_color='C6EFCE', fill_type='solid')
                    cell.font = Font(color='006100', bold=True, size=12)
                elif value == 'N':
                    cell.fill = PatternFill(start_color='FFC7CE', fill_type='solid')
                    cell.font = Font(color='9C0006', bold=True, size=12)
                elif value == 'With Changes':
                    cell.fill = PatternFill(start_color='FFEB9C', fill_type='solid')
                    cell.font = Font(color='9C5700', bold=True, size=11)
    
    # Column widths
    ws.column_dimensions['A'].width = 65  # Feature
    ws.column_dimensions['B'].width = 15  # Approved
    ws.column_dimensions['C'].width = 50  # Notes
    ws.column_dimensions['D'].width = 10  # UUID
    
    ws.freeze_panes = 'A2'
    
    # Stats Summary
    ws['F1'] = 'SUMMARY'
    ws['F1'].font = Font(bold=True, color='FFFFFF', size=14)
    ws['F1'].fill = PatternFill(start_color='4472C4', fill_type='solid')
    
    ws['F2'] = 'Total'
    ws['G2'] = '=COUNTA(A:A)-1'
    ws['F3'] = 'Approved (Y)'
    ws['G3'] = '=COUNTIF(B:B,"Y")'
    ws['F4'] = 'Rejected (N)'
    ws['G4'] = '=COUNTIF(B:B,"N")'
    ws['F5'] = 'With Changes'
    ws['G5'] = '=COUNTIF(B:B,"With Changes")'
    
    for row in range(2, 6):
        ws[f'F{row}'].font = Font(bold=True)
        ws[f'F{row}'].fill = PatternFill(start_color='E7E6E6', fill_type='solid')
    
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    
    # Instructions
    ws.insert_rows(1)
    ws['A1'] = '📋 KEYBOARD: Tab → Feature → Approved (Y/N/With Changes) → Notes → Tab to next row | Arrow keys also work'
    ws['A1'].font = Font(bold=True, size=11, color='1F4E78')
    ws['A1'].fill = PatternFill(start_color='D9E1F2', fill_type='solid')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    filepath = os.path.join(output_dir, 'REQUIREMENTS-SIMPLE.xlsx')
    wb.save(filepath)
    return filepath

def create_advanced_template(output_dir):
    """Create advanced requirements template with feature groups, tags, and versions"""
    wb = Workbook()
    wb.remove(wb.active)
    
    header_fill = PatternFill(start_color='2F5496', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Main Requirements Sheet
    ws = wb.create_sheet('Requirements')
    
    headers = ['Feature Group', 'Tags', 'Type', 'Requirement', 'Must/Nice', 'v1', 'v2', 'v3', 
                'Decision', 'Status', 'User Feedback', 'Revision Notes', 'UUID']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sample data
    sample_data = [
        ['Core Features', 'UX, MVP', 'Functional', 'Sample feature 1', 'MUST', 'X', '', '', 
         'Accepted', 'In Progress', '✓ Looks good', '', gen_uuid()],
        ['Core Features', 'Backend, API', 'Functional', 'Sample feature 2', 'Nice', '', 'X', '', 
         'Accepted with Changes', 'Not Started', 'Needs minor tweaks', 'Updated scope', gen_uuid()],
        ['Quality', 'Performance', 'Non-Functional', 'Load time < 2s', 'MUST', 'X', '', '', 
         'Accepted', 'Not Started', '✓ Standard requirement', '', gen_uuid()],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Color code Must/Nice
            if col_idx == 5:
                if value == 'MUST':
                    cell.fill = PatternFill(start_color='FFE699', fill_type='solid')
                    cell.font = Font(bold=True)
                elif value == 'Nice':
                    cell.fill = PatternFill(start_color='E2EFDA', fill_type='solid')
            
            # Color code Decision
            if col_idx == 9:
                if value == 'Accepted':
                    cell.fill = PatternFill(start_color='C6EFCE', fill_type='solid')
                    cell.font = Font(color='006100', bold=True)
                elif value == 'Accepted with Changes':
                    cell.fill = PatternFill(start_color='FFEB9C', fill_type='solid')
                    cell.font = Font(color='9C5700', bold=True)
                elif value == 'Revise':
                    cell.fill = PatternFill(start_color='FFCCCC', fill_type='solid')
                    cell.font = Font(color='CC0000', bold=True)
                elif value == 'Kill':
                    cell.fill = PatternFill(start_color='C0C0C0', fill_type='solid')
                    cell.font = Font(color='666666', strikethrough=True)
            
            if col_idx == 1:  # Bold feature groups
                cell.font = Font(bold=True)
    
    # Column widths
    widths = [20, 25, 15, 50, 12, 5, 5, 5, 20, 15, 40, 40, 10]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64+idx)].width = width
    
    ws.freeze_panes = 'A2'
    
    # Instructions
    ws.insert_rows(1)
    ws['A1'] = '⌨️ KEYBOARD: Tab left-to-right | Enter moves down | Arrow keys navigate | UUID auto-generated at end'
    ws['A1'].font = Font(bold=True, size=11, color='1F4E78')
    ws['A1'].fill = PatternFill(start_color='D9E1F2', fill_type='solid')
    ws.merge_cells('A1:M1')
    ws.row_dimensions[1].height = 30
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Summary Stats Sheet
    ws_stats = wb.create_sheet('Summary Stats')
    ws_stats['A1'] = 'REQUIREMENT STATISTICS'
    ws_stats['A1'].font = Font(bold=True, size=16)
    
    stats_data = [
        ['', ''],
        ['OVERALL', 'COUNT'],
        ['Total Requirements', '=COUNTA(Requirements!D:D)-1'],
        ['MUST Have', '=COUNTIF(Requirements!E:E,"MUST")'],
        ['Nice to Have', '=COUNTIF(Requirements!E:E,"Nice")'],
        ['', ''],
        ['VERSION BREAKDOWN', 'COUNT'],
        ['v1 Requirements', '=COUNTIF(Requirements!F:F,"X")'],
        ['v2 Requirements', '=COUNTIF(Requirements!G:G,"X")'],
        ['v3 Requirements', '=COUNTIF(Requirements!H:H,"X")'],
        ['', ''],
        ['DECISIONS', 'COUNT'],
        ['Accepted', '=COUNTIF(Requirements!I:I,"Accepted")'],
        ['Rejected', '=COUNTIF(Requirements!I:I,"Kill")'],
        ['Pending', '=COUNTIF(Requirements!I:I,"Pending")'],
    ]
    
    for row_idx, row_data in enumerate(stats_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_stats.cell(row_idx, col_idx, value)
            if row_idx in [2, 7, 12]:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color='4472C4', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
    
    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 15
    
    filepath = os.path.join(output_dir, 'REQUIREMENTS-ADVANCED.xlsx')
    wb.save(filepath)
    return filepath

def main():
    """Main function to generate both templates"""
    output_dir = sys.argv[1] if len(sys.argv) > 1 else '.'
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    print('Generating requirements templates...')
    
    simple_path = create_simple_template(output_dir)
    print(f'✓ Created {simple_path}')
    print('  • 4 columns: Feature, Approved (Y/N/With Changes), Notes, UUID')
    print('  • Keyboard optimized')
    print('  • Auto-calculating stats')
    
    advanced_path = create_advanced_template(output_dir)
    print(f'✓ Created {advanced_path}')
    print('  • 13 columns: Feature Groups, Tags, Type, Versions, Decisions, etc.')
    print('  • UUID at end')
    print('  • Summary stats sheet')
    
    print('\n✅ Templates ready to use!')

if __name__ == '__main__':
    main()
